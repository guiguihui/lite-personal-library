"""XLSX 解析器 — 基于 openpyxl。

逐 sheet 提取单元格文本,每行作为一个 chunk。
页码 = sheet 序号,page_label = sheet 名,行号 = Excel 行号(1-based)。
"""

from __future__ import annotations

import logging
from pathlib import Path

from app.fileparse.base import BaseParser, ParseChunk

logger = logging.getLogger("lqd.fileparse.xlsx")


class XlsxParser(BaseParser):
    extensions = [".xlsx"]

    def parse(self, file_path: Path) -> list[ParseChunk]:
        from openpyxl import load_workbook

        logger.debug("XLSX 解析开始: %s", file_path.name)
        chunks: list[ParseChunk] = []
        try:
            wb = load_workbook(str(file_path), read_only=True, data_only=True)
        except Exception as exc:
            logger.error("XLSX 打开失败 %s: %s", file_path, exc)
            raise

        try:
            sheet_count = len(wb.sheetnames)
            logger.info("XLSX %s: 共 %d 个工作表: %s",
                        file_path.name, sheet_count, wb.sheetnames)

            for sheet_idx, sheet_name in enumerate(wb.sheetnames, 1):
                ws = wb[sheet_name]
                row_count = 0
                current_lines: list[str] = []
                current_start_row = 0
                target_chars = 500

                for row in ws.iter_rows(values_only=True):
                    row_count += 1
                    # 跳过全空行
                    if all(v is None or str(v).strip() == "" for v in row):
                        continue

                    row_texts = []
                    for cell in row:
                        if cell is not None:
                            cell_text = str(cell).strip()
                            if cell_text:
                                row_texts.append(cell_text)

                    if not row_texts:
                        continue

                    if not current_lines:
                        current_start_row = row_count
                    current_lines.append(" | ".join(row_texts))
                    total_chars = sum(len(l) for l in current_lines)

                    if total_chars >= target_chars:
                        chunk_text = "\n".join(current_lines)
                        chunks.append(ParseChunk(
                            text=chunk_text,
                            page=sheet_idx,
                            page_label=sheet_name,
                            line_start=current_start_row,
                            line_end=row_count,
                            metadata={"sheet": sheet_name},
                        ))
                        current_lines = []
                        current_start_row = row_count + 1

                # 剩余行
                if current_lines:
                    chunk_text = "\n".join(current_lines)
                    chunks.append(ParseChunk(
                        text=chunk_text,
                        page=sheet_idx,
                        page_label=sheet_name,
                        line_start=current_start_row,
                        line_end=row_count,
                        metadata={"sheet": sheet_name},
                    ))

                logger.debug("XLSX %s sheet '%s': %d 行, 已生成切片",
                             file_path.name, sheet_name, row_count)
        finally:
            wb.close()

        logger.info("XLSX 解析完成: %s, %d sheets, %d 切片",
                    file_path.name, len(wb.sheetnames), len(chunks))
        return chunks
